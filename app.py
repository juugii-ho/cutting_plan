#!/usr/bin/env python3
"""
H-Beam 철판 절단 최적화 웹 애플리케이션 (PDF 포함 완전판)
Streamlit Cloud 무료 배포 가능
"""

import streamlit as st
import pandas as pd
import io
import math
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# PDF용
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.lib import colors

# ============================================================
# 제약조건 상수
# ============================================================
MAX_LENGTH = 22000
MIN_LENGTH = 4000
MAX_WIDTH = 4200
MIN_WIDTH = 1500
MAX_WEIGHT = 18000

PREFERRED_LENGTH_MAX = 16000
PREFERRED_LENGTH_MIN = 8000
PREFERRED_WIDTH_MAX = 3000
PREFERRED_WIDTH_MIN = 2100

NO_MERGE_LENGTH = 8000
STEEL_DENSITY = 7.85e-6
LOSS_THRESHOLD = 0.10

# ============================================================
# 데이터 클래스
# ============================================================
@dataclass
class PlateItem:
    source_no: int
    plate_type: int
    material: str
    thickness: float
    length: float
    width: float
    quantity: int = 1

@dataclass
class PlacedPlate:
    source_no: int
    plate_type: int
    material: str
    thickness: float
    length: float
    width: float
    x: float
    y: float

@dataclass
class MasterPlate:
    seq: int
    material: str
    thickness: float
    length: float
    width: float
    placed_plates: List[PlacedPlate] = field(default_factory=list)
    
    @property
    def weight(self) -> float:
        return self.length * self.width * self.thickness * STEEL_DENSITY
    
    @property
    def usage_rate(self) -> float:
        if self.length * self.width == 0:
            return 0
        used = sum(p.length * p.width for p in self.placed_plates)
        return used / (self.length * self.width)
    
    @property
    def plate_count(self) -> int:
        return len(self.placed_plates)

# ============================================================
# 유틸리티 함수
# ============================================================
def calculate_weight(thickness: float, width: float, length: float) -> float:
    return length * width * thickness * STEEL_DENSITY

def get_plate_type_name(plate_type: int) -> str:
    return {1: "W", 2: "UF", 3: "LF"}.get(plate_type, "?")

# ============================================================
# H빔 주문 → 개별 철판 변환
# ============================================================
def convert_hbeam_to_plates(orders: List[Dict]) -> List[PlateItem]:
    plates = []
    for order in orders:
        source_no = order['no']
        material = order['material']
        qty = order['quantity']
        
        h = order['H']
        b = order['B']
        tw = order['tw']
        tf = order['tf']
        length = order['length']
        
        web_height = h - 2 * tf
        plates.append(PlateItem(
            source_no=source_no, plate_type=1, material=material,
            thickness=tw, length=length, width=web_height, quantity=qty
        ))
        
        plates.append(PlateItem(
            source_no=source_no, plate_type=2, material=material,
            thickness=tf, length=length, width=b, quantity=qty
        ))
        
        plates.append(PlateItem(
            source_no=source_no, plate_type=3, material=material,
            thickness=tf, length=length, width=b, quantity=qty
        ))
    
    return plates

# ============================================================
# BinPacker 클래스
# ============================================================
class BinPacker:
    def __init__(self, material: str, thickness: float, allow_rotation: bool = False):
        self.material = material
        self.thickness = thickness
        self.allow_rotation = allow_rotation
        self.bins: List[Dict] = []
    
    def pack(self, plates: List[Tuple[int, int, float, float]]) -> List[MasterPlate]:
        long_plates = [p for p in plates if p[2] >= NO_MERGE_LENGTH]
        short_plates = [p for p in plates if p[2] < NO_MERGE_LENGTH]
        
        results = []
        
        if long_plates:
            groups: Dict[Tuple[float, float], List] = defaultdict(list)
            for p in long_plates:
                key = (p[2], p[3])
                groups[key].append(p)
            
            for (length, width), group_plates in groups.items():
                bins = self._pack_uniform_group_optimized(group_plates, length)
                for bin_info in bins:
                    results.append(self._create_master_plate(bin_info, length))
        
        if short_plates:
            groups: Dict[Tuple[float, float], List] = defaultdict(list)
            for p in short_plates:
                key = (p[2], p[3])
                groups[key].append(p)
            
            for (length, width), group_plates in groups.items():
                bins = self._pack_uniform_group_2d_optimized(group_plates, length)
                for bin_info in bins:
                    mp_length = bin_info.get('total_length', length)
                    results.append(self._create_master_plate(bin_info, mp_length))
        
        return results
    
    def _pack_uniform_group_optimized(self, plates: List, length: float) -> List[Dict]:
        if not plates:
            return []
        
        plate_width = plates[0][3]
        total_plates = len(plates)
        
        max_rows = int(MAX_LENGTH / length) if length > 0 else 1
        
        if max_rows > 1:
            best_layout = self._find_optimal_2d_layout(total_plates, length, plate_width)
            
            if best_layout:
                rows, cols_per_bin = best_layout
                bins = []
                plate_idx = 0
                
                while plate_idx < len(plates):
                    new_bin = {'width_used': 0, 'plates': [], 'total_length': rows * length}
                    
                    for row in range(rows):
                        for col in range(cols_per_bin):
                            if plate_idx >= len(plates):
                                break
                            
                            source_no, plate_type, pl_length, pl_width = plates[plate_idx]
                            x = col * pl_width
                            y = row * pl_length
                            
                            new_bin['plates'].append(
                                PlacedPlate(source_no, plate_type, self.material,
                                           self.thickness, pl_length, pl_width, x, y)
                            )
                            plate_idx += 1
                    
                    if new_bin['plates']:
                        max_col = max(p.x for p in new_bin['plates']) // plate_width + 1 if new_bin['plates'] else 1
                        new_bin['width_used'] = max_col * plate_width
                        bins.append(new_bin)
                
                return bins
        
        max_cols_absolute = int(MAX_WIDTH / plate_width) if plate_width > 0 else 1
        max_cols_preferred = int(PREFERRED_WIDTH_MAX / plate_width) if plate_width > 0 else 1
        
        if max_cols_preferred < 1:
            max_cols_preferred = 1
        if max_cols_absolute < 1:
            max_cols_absolute = 1
        
        bins = self._distribute_plates(plates, plate_width, max_cols_preferred)
        
        for _ in range(20):
            if len(bins) < 2:
                break
            
            last_bin = bins[-1]
            last_cols = len(last_bin['plates'])
            last_width = last_cols * plate_width
            
            need_rebalance = False
            
            if last_width < MIN_WIDTH:
                need_rebalance = True
            elif last_width < PREFERRED_WIDTH_MIN:
                loss = 1 - (last_width / PREFERRED_WIDTH_MIN)
                if loss > LOSS_THRESHOLD:
                    need_rebalance = True
            
            if not need_rebalance:
                break
            
            prev_bin = bins[-2]
            combined_plates = prev_bin['plates'] + last_bin['plates']
            total_cols = len(combined_plates)
            
            best_distribution = self._find_best_distribution(
                total_cols, plate_width, length, max_cols_absolute
            )
            
            if best_distribution is None:
                break
            
            new_bins = self._create_bins_from_distribution(
                combined_plates, best_distribution, plate_width
            )
            
            old_area = self._calc_total_area(bins[-2:], plate_width, length)
            new_area = self._calc_total_area(new_bins, plate_width, length)
            
            if new_area < old_area:
                bins = bins[:-2] + new_bins
            else:
                break
        
        return bins
    
    def _find_optimal_2d_layout(self, n: int, plate_length: float, plate_width: float) -> Optional[Tuple[int, int]]:
        best_layout = None
        best_area = float('inf')
        best_preferred = False
        
        max_rows = min(int(MAX_LENGTH / plate_length), n) if plate_length > 0 else n
        max_cols = min(int(MAX_WIDTH / plate_width), n) if plate_width > 0 else n
        
        for rows in range(1, max_rows + 1):
            cols_needed = (n + rows - 1) // rows
            
            if cols_needed > max_cols:
                continue
            
            total_length = rows * plate_length
            total_width = cols_needed * plate_width
            
            if total_length > MAX_LENGTH or total_width > MAX_WIDTH:
                continue
            if total_length < MIN_LENGTH:
                total_length = MIN_LENGTH
            if total_width < MIN_WIDTH:
                total_width = MIN_WIDTH
            
            if calculate_weight(self.thickness, total_width, total_length) > MAX_WEIGHT:
                continue
            
            total_area = total_length * total_width
            
            is_preferred = (PREFERRED_LENGTH_MIN <= total_length <= PREFERRED_LENGTH_MAX and
                           PREFERRED_WIDTH_MIN <= total_width <= PREFERRED_WIDTH_MAX)
            
            if total_area < best_area or (total_area == best_area and is_preferred and not best_preferred):
                best_area = total_area
                best_layout = (rows, cols_needed)
                best_preferred = is_preferred
        
        return best_layout
    
    def _find_best_distribution(self, total_cols: int, plate_width: float, 
                                 length: float, max_cols: int) -> Optional[List[int]]:
        best_option = None
        best_area = float('inf')
        best_preferred_score = -1
        
        if total_cols * plate_width <= MAX_WIDTH:
            width = total_cols * plate_width
            effective_width = max(width, MIN_WIDTH)
            area = effective_width * length
            
            if calculate_weight(self.thickness, effective_width, length) <= MAX_WEIGHT:
                preferred_score = 1 if PREFERRED_WIDTH_MIN <= width <= PREFERRED_WIDTH_MAX else 0
                
                if area < best_area or (area == best_area and preferred_score > best_preferred_score):
                    best_area = area
                    best_option = [total_cols]
                    best_preferred_score = preferred_score
        
        for cols1 in range(1, total_cols):
            cols2 = total_cols - cols1
            
            width1 = cols1 * plate_width
            width2 = cols2 * plate_width
            
            if width1 > MAX_WIDTH or width2 > MAX_WIDTH:
                continue
            
            eff_width1 = max(width1, MIN_WIDTH)
            eff_width2 = max(width2, MIN_WIDTH)
            
            if calculate_weight(self.thickness, eff_width1, length) > MAX_WEIGHT:
                continue
            if calculate_weight(self.thickness, eff_width2, length) > MAX_WEIGHT:
                continue
            
            area = eff_width1 * length + eff_width2 * length
            
            pref1 = 1 if PREFERRED_WIDTH_MIN <= width1 <= PREFERRED_WIDTH_MAX else 0
            pref2 = 1 if PREFERRED_WIDTH_MIN <= width2 <= PREFERRED_WIDTH_MAX else 0
            preferred_score = pref1 + pref2
            
            if area < best_area or (area == best_area and preferred_score > best_preferred_score):
                best_area = area
                best_option = [cols1, cols2]
                best_preferred_score = preferred_score
        
        return best_option
    
    def _distribute_plates(self, plates: List, plate_width: float, max_cols: int) -> List[Dict]:
        total_plates = len(plates)
        
        num_bins = (total_plates + max_cols - 1) // max_cols
        if num_bins < 1:
            num_bins = 1
        
        base_count = total_plates // num_bins
        extra = total_plates % num_bins
        
        bins = []
        plate_idx = 0
        
        for i in range(num_bins):
            count = base_count + (1 if i < extra else 0)
            new_bin = {'width_used': 0, 'plates': []}
            
            for col in range(count):
                if plate_idx >= len(plates):
                    break
                
                source_no, plate_type, pl_length, pl_width = plates[plate_idx]
                x = col * pl_width
                
                new_bin['plates'].append(
                    PlacedPlate(source_no, plate_type, self.material,
                               self.thickness, pl_length, pl_width, x, 0)
                )
                plate_idx += 1
            
            new_bin['width_used'] = count * plate_width
            bins.append(new_bin)
        
        return bins
    
    def _create_bins_from_distribution(self, plates: List, distribution: List[int], 
                                        plate_width: float) -> List[Dict]:
        bins = []
        plate_idx = 0
        
        for count in distribution:
            new_bin = {'width_used': 0, 'plates': []}
            
            for col in range(count):
                if plate_idx >= len(plates):
                    break
                
                plate = plates[plate_idx]
                if isinstance(plate, PlacedPlate):
                    plate.x = col * plate_width
                    new_bin['plates'].append(plate)
                else:
                    source_no, plate_type, pl_length, pl_width = plate
                    new_bin['plates'].append(
                        PlacedPlate(source_no, plate_type, self.material,
                                   self.thickness, pl_length, pl_width, col * plate_width, 0)
                    )
                plate_idx += 1
            
            new_bin['width_used'] = count * plate_width
            bins.append(new_bin)
        
        return bins
    
    def _calc_total_area(self, bins: List[Dict], plate_width: float, length: float) -> float:
        total = 0
        for b in bins:
            width = max(b['width_used'], MIN_WIDTH)
            total += width * length
        return total
    
    def _pack_uniform_group_2d_optimized(self, plates: List, length: float) -> List[Dict]:
        if not plates:
            return []
        
        plate_width = plates[0][3]
        plate_length = length
        total_plates = len(plates)
        
        best_layout = self._find_optimal_2d_layout(total_plates, plate_length, plate_width)
        
        if best_layout is None:
            rows_per_col = max(1, int(PREFERRED_LENGTH_MAX / plate_length))
            max_cols = max(1, int(PREFERRED_WIDTH_MAX / plate_width))
            best_layout = (rows_per_col, max_cols)
        
        rows_per_col, max_cols_per_bin = best_layout
        
        bins = self._distribute_2d(plates, plate_width, plate_length, 
                                   rows_per_col, max_cols_per_bin)
        
        master_length = rows_per_col * plate_length
        max_cols_absolute = int(MAX_WIDTH / plate_width) if plate_width > 0 else 1
        
        for _ in range(20):
            if len(bins) < 2:
                break
            
            last_bin = bins[-1]
            last_width = last_bin['width_used']
            
            need_rebalance = False
            if last_width < MIN_WIDTH:
                need_rebalance = True
            elif last_width < PREFERRED_WIDTH_MIN:
                loss = 1 - (last_width / PREFERRED_WIDTH_MIN)
                if loss > LOSS_THRESHOLD:
                    need_rebalance = True
            
            if not need_rebalance:
                break
            
            prev_bin = bins[-2]
            combined_plates = prev_bin['plates'] + last_bin['plates']
            total_cols_combined = (len(combined_plates) + rows_per_col - 1) // rows_per_col
            
            best_distribution = self._find_best_distribution(
                total_cols_combined, plate_width, master_length, max_cols_absolute
            )
            
            if best_distribution is None:
                break
            
            new_bins = self._create_bins_2d_from_distribution(
                combined_plates, best_distribution, plate_width, plate_length, rows_per_col
            )
            
            old_area = self._calc_total_area(bins[-2:], plate_width, master_length)
            new_area = self._calc_total_area(new_bins, plate_width, master_length)
            
            if new_area < old_area:
                bins = bins[:-2] + new_bins
            else:
                break
        
        return bins
    
    def _distribute_2d(self, plates: List, plate_width: float, plate_length: float,
                       rows_per_col: int, max_cols: int) -> List[Dict]:
        total_plates = len(plates)
        total_cols = (total_plates + rows_per_col - 1) // rows_per_col
        
        num_bins = (total_cols + max_cols - 1) // max_cols
        if num_bins < 1:
            num_bins = 1
        
        base_cols = total_cols // num_bins
        extra_cols = total_cols % num_bins
        
        bins = []
        plate_idx = 0
        
        for i in range(num_bins):
            num_cols = base_cols + (1 if i < extra_cols else 0)
            new_bin = {'width_used': 0, 'plates': [], 'total_length': rows_per_col * plate_length}
            
            for col in range(num_cols):
                col_x = col * plate_width
                
                for row in range(rows_per_col):
                    if plate_idx >= len(plates):
                        break
                    
                    source_no, plate_type, pl_length, pl_width = plates[plate_idx]
                    y = row * pl_length
                    
                    new_bin['plates'].append(
                        PlacedPlate(source_no, plate_type, self.material,
                                   self.thickness, pl_length, pl_width, col_x, y)
                    )
                    plate_idx += 1
            
            new_bin['width_used'] = num_cols * plate_width
            bins.append(new_bin)
        
        return bins
    
    def _create_bins_2d_from_distribution(self, plates: List, distribution: List[int],
                                           plate_width: float, plate_length: float,
                                           rows_per_col: int) -> List[Dict]:
        bins = []
        plate_idx = 0
        
        for num_cols in distribution:
            new_bin = {'width_used': 0, 'plates': [], 'total_length': rows_per_col * plate_length}
            
            for col in range(num_cols):
                col_x = col * plate_width
                
                for row in range(rows_per_col):
                    if plate_idx >= len(plates):
                        break
                    
                    plate = plates[plate_idx]
                    if isinstance(plate, PlacedPlate):
                        plate.x = col_x
                        plate.y = row * plate_length
                        new_bin['plates'].append(plate)
                    else:
                        source_no, plate_type, pl_length, pl_width = plate
                        new_bin['plates'].append(
                            PlacedPlate(source_no, plate_type, self.material,
                                       self.thickness, pl_length, pl_width, col_x, row * pl_length)
                        )
                    plate_idx += 1
            
            new_bin['width_used'] = num_cols * plate_width
            bins.append(new_bin)
        
        return bins
    
    def _create_master_plate(self, bin_info: Dict, length: float) -> MasterPlate:
        width = bin_info.get('width_used', bin_info.get('max_width', MIN_WIDTH))
        width = max(width, MIN_WIDTH)
        length = max(bin_info.get('total_length', length), MIN_LENGTH)
        
        plates = bin_info.get('plates', [])
        if not plates and 'shelves' in bin_info:
            plates = [p for shelf in bin_info['shelves'] for p in shelf['plates']]
        
        return MasterPlate(
            seq=0,
            material=self.material,
            thickness=self.thickness,
            length=length,
            width=width,
            placed_plates=plates
        )

# ============================================================
# 메인 최적화 함수
# ============================================================
def optimize_cutting(plates: List[PlateItem], allow_rotation: bool = False) -> List[MasterPlate]:
    groups: Dict[Tuple, List] = defaultdict(list)
    
    for plate in plates:
        key = (plate.material, plate.thickness, plate.plate_type)
        for _ in range(plate.quantity):
            groups[key].append((plate.source_no, plate.plate_type, plate.length, plate.width))
    
    all_results = []
    
    for (material, thickness, plate_type), plate_list in groups.items():
        packer = BinPacker(material, thickness, allow_rotation)
        results = packer.pack(plate_list)
        all_results.extend(results)
    
    all_results = merge_low_usage_bins(all_results)
    
    all_results.sort(key=lambda mp: (mp.material, mp.thickness, -mp.length, -mp.width))
    for i, mp in enumerate(all_results, 1):
        mp.seq = i
    
    return all_results


def merge_low_usage_bins(master_plates: List[MasterPlate], threshold: float = 0.85) -> List[MasterPlate]:
    good_bins = []
    low_usage_bins: Dict[Tuple[str, float], List[MasterPlate]] = defaultdict(list)
    
    for mp in master_plates:
        if mp.usage_rate < threshold:
            key = (mp.material, mp.thickness)
            low_usage_bins[key].append(mp)
        else:
            good_bins.append(mp)
    
    merged_bins = []
    
    for (material, thickness), bins in low_usage_bins.items():
        if len(bins) < 2:
            merged_bins.extend(bins)
            continue
        
        all_plates = []
        for mp in bins:
            for p in mp.placed_plates:
                all_plates.append((p.source_no, p.plate_type, p.length, p.width))
        
        if not all_plates:
            continue
        
        widths = set(p[3] for p in all_plates)
        lengths = set(p[2] for p in all_plates)
        
        if len(widths) == 1 and len(lengths) == 1:
            plate_width = all_plates[0][3]
            plate_length = all_plates[0][2]
            total_plates = len(all_plates)
            
            best_layout = find_optimal_layout_for_merge(
                total_plates, plate_length, plate_width, thickness
            )
            
            if best_layout:
                rows, cols = best_layout
                new_bins = create_merged_bins(
                    all_plates, material, thickness, 
                    plate_length, plate_width, rows, cols
                )
                merged_bins.extend(new_bins)
            else:
                merged_bins.extend(bins)
        else:
            merged_bins.extend(bins)
    
    return good_bins + merged_bins


def find_optimal_layout_for_merge(n: int, plate_length: float, plate_width: float, 
                                   thickness: float) -> Optional[Tuple[int, int]]:
    best_layout = None
    best_area = float('inf')
    
    max_rows = min(int(MAX_LENGTH / plate_length), n) if plate_length > 0 else n
    max_cols = min(int(MAX_WIDTH / plate_width), n) if plate_width > 0 else n
    
    for rows in range(1, max_rows + 1):
        cols = (n + rows - 1) // rows
        
        if cols > max_cols:
            continue
        
        total_length = max(rows * plate_length, MIN_LENGTH)
        total_width = max(cols * plate_width, MIN_WIDTH)
        
        if total_length > MAX_LENGTH or total_width > MAX_WIDTH:
            continue
        
        if calculate_weight(thickness, total_width, total_length) > MAX_WEIGHT:
            continue
        
        area = total_length * total_width
        
        if area < best_area:
            best_area = area
            best_layout = (rows, cols)
    
    return best_layout


def create_merged_bins(plates: List, material: str, thickness: float,
                       plate_length: float, plate_width: float,
                       rows: int, cols: int) -> List[MasterPlate]:
    results = []
    plate_idx = 0
    
    total_length = rows * plate_length
    total_width = cols * plate_width
    
    placed = []
    for row in range(rows):
        for col in range(cols):
            if plate_idx >= len(plates):
                break
            
            source_no, plate_type, pl_length, pl_width = plates[plate_idx]
            x = col * plate_width
            y = row * plate_length
            
            placed.append(PlacedPlate(
                source_no, plate_type, material, thickness,
                pl_length, pl_width, x, y
            ))
            plate_idx += 1
    
    if placed:
        mp = MasterPlate(
            seq=0,
            material=material,
            thickness=thickness,
            length=max(total_length, MIN_LENGTH),
            width=max(total_width, MIN_WIDTH),
            placed_plates=placed
        )
        results.append(mp)
    
    while plate_idx < len(plates):
        remaining = len(plates) - plate_idx
        sub_layout = find_optimal_layout_for_merge(remaining, plate_length, plate_width, thickness)
        
        if sub_layout is None:
            sub_layout = (1, remaining)
        
        sub_rows, sub_cols = sub_layout
        placed = []
        
        for row in range(sub_rows):
            for col in range(sub_cols):
                if plate_idx >= len(plates):
                    break
                
                source_no, plate_type, pl_length, pl_width = plates[plate_idx]
                x = col * plate_width
                y = row * plate_length
                
                placed.append(PlacedPlate(
                    source_no, plate_type, material, thickness,
                    pl_length, pl_width, x, y
                ))
                plate_idx += 1
        
        if placed:
            sub_length = max(sub_rows * plate_length, MIN_LENGTH)
            sub_width = max(sub_cols * plate_width, MIN_WIDTH)
            
            mp = MasterPlate(
                seq=0,
                material=material,
                thickness=thickness,
                length=sub_length,
                width=sub_width,
                placed_plates=placed
            )
            results.append(mp)
    
    return results


# ============================================================
# Excel 저장 함수들
# ============================================================
def save_result_formatted(master_plates: List[MasterPlate], orders: List[Dict]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cutting Result"
    
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    headers_row1 = ["NO", "순위", "BOM\n재질", "MARK", "SIZE", "", "", "", "", "규격", "길이", "Q'TY", "Total\nLength", "단중", "견적중량"]
    headers_row2 = ["", "", "", "", "SEC", "H1", "B", "T1", "T2", "", "", "", "", "", ""]
    
    for col, header in enumerate(headers_row1, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align
    
    for col, header in enumerate(headers_row2, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align
    
    ws.merge_cells('E1:I1')
    for col in [1, 2, 3, 4, 10, 11, 12, 13, 14, 15]:
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    
    order_dict = {o['no']: o for o in orders}
    
    used_orders = defaultdict(lambda: {'qty': 0, 'length': 0})
    for mp in master_plates:
        for plate in mp.placed_plates:
            used_orders[plate.source_no]['qty'] += 1
            used_orders[plate.source_no]['length'] += plate.length
    
    row_num = 3
    total_qty = 0
    total_length = 0
    total_weight = 0
    
    for no, usage in sorted(used_orders.items()):
        if no not in order_dict:
            continue
        
        order = order_dict[no]
        qty = usage['qty'] // 3
        if qty == 0:
            qty = usage['qty']
        
        length = order['length']
        total_len = qty * length
        
        h, b, tw, tf = order['H'], order['B'], order['tw'], order['tf']
        area = (h - 2*tf) * tw + 2 * b * tf
        unit_weight = area * STEEL_DENSITY * 1000
        est_weight = unit_weight * total_len / 1000
        
        mark = f"UJ-{no:03d}"
        spec = f"BH-{int(h)}x{int(b)}x{int(tw)}x{int(tf)}"
        
        ws.cell(row=row_num, column=1, value=no)
        ws.cell(row=row_num, column=2, value="")
        ws.cell(row=row_num, column=3, value=order['material'])
        ws.cell(row=row_num, column=4, value=mark)
        ws.cell(row=row_num, column=5, value="BH")
        ws.cell(row=row_num, column=6, value=int(h))
        ws.cell(row=row_num, column=7, value=int(b))
        ws.cell(row=row_num, column=8, value=int(tw))
        ws.cell(row=row_num, column=9, value=int(tf))
        ws.cell(row=row_num, column=10, value=spec)
        ws.cell(row=row_num, column=11, value=int(length))
        ws.cell(row=row_num, column=12, value=qty)
        ws.cell(row=row_num, column=13, value=int(total_len))
        ws.cell(row=row_num, column=14, value=round(unit_weight, 3))
        ws.cell(row=row_num, column=15, value=round(est_weight, 1))
        
        for col in range(1, 16):
            ws.cell(row=row_num, column=col).border = thin_border
            ws.cell(row=row_num, column=col).alignment = center_align
        
        total_qty += qty
        total_length += total_len
        total_weight += est_weight
        row_num += 1
    
    ws.cell(row=row_num, column=1, value="합       계")
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=11)
    ws.cell(row=row_num, column=12, value=total_qty)
    ws.cell(row=row_num, column=13, value=int(total_length))
    ws.cell(row=row_num, column=14, value=round(total_weight / total_length * 1000, 3) if total_length > 0 else 0)
    ws.cell(row=row_num, column=15, value=round(total_weight, 1))
    
    for col in range(1, 16):
        cell = ws.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.alignment = center_align
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    widths = [6, 6, 12, 10, 6, 6, 6, 6, 6, 22, 10, 8, 12, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_summary_excel(master_plates: List[MasterPlate]) -> io.BytesIO:
    data = []
    for mp in master_plates:
        details = "; ".join([f"#{p.source_no}({get_plate_type_name(p.plate_type)})" 
                            for p in mp.placed_plates[:5]])
        if len(mp.placed_plates) > 5:
            details += f"... +{len(mp.placed_plates)-5}"
        
        data.append({
            'No': mp.seq,
            'Material': mp.material,
            'Thickness(mm)': mp.thickness,
            'Length(mm)': mp.length,
            'Width(mm)': mp.width,
            'Weight(kg)': round(mp.weight, 1),
            'Plates': mp.plate_count,
            'Usage%': round(mp.usage_rate * 100, 1),
            'Details': details
        })
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    return output


# ============================================================
# PDF 생성
# ============================================================
def get_tick_step(size: float) -> int:
    if size <= 2000: return 500
    elif size <= 5000: return 1000
    elif size <= 10000: return 2000
    elif size <= 20000: return 5000
    else: return 10000

def generate_pdf(master_plates: List[MasterPlate]) -> io.BytesIO:
    output = io.BytesIO()
    c = pdf_canvas.Canvas(output, pagesize=landscape(A4))
    page_width, page_height = landscape(A4)
    
    margin = 30
    
    # 색상 매핑
    material_thickness_set = set()
    for mp in master_plates:
        material_thickness_set.add((mp.material, mp.thickness))
    
    color_palette = [
        colors.Color(0.98, 0.6, 0.6),
        colors.Color(0.6, 0.98, 0.6),
        colors.Color(0.6, 0.6, 0.98),
        colors.Color(0.98, 0.98, 0.6),
        colors.Color(0.98, 0.6, 0.98),
        colors.Color(0.6, 0.98, 0.98),
        colors.Color(0.98, 0.8, 0.6),
        colors.Color(0.8, 0.6, 0.98),
        colors.Color(0.6, 0.98, 0.8),
        colors.Color(0.98, 0.6, 0.8),
        colors.Color(0.8, 0.98, 0.6),
        colors.Color(0.6, 0.8, 0.98),
    ]
    
    material_thickness_colors = {}
    for i, (mat, thk) in enumerate(sorted(material_thickness_set)):
        material_thickness_colors[(mat, thk)] = color_palette[i % len(color_palette)]
    
    # 패턴 그룹화
    def get_pattern_key(mp):
        plate_info = tuple(sorted([(p.length, p.width, p.x, p.y) for p in mp.placed_plates]))
        return (mp.material, mp.thickness, mp.length, mp.width, plate_info)
    
    pattern_groups = defaultdict(list)
    for mp in master_plates:
        key = get_pattern_key(mp)
        pattern_groups[key].append(mp)
    
    unique_patterns = []
    for key, mps in pattern_groups.items():
        unique_patterns.append((mps[0], mps[0].seq, mps[-1].seq, len(mps)))
    unique_patterns.sort(key=lambda x: x[1])
    
    # 요약 페이지
    c.setFont("Helvetica-Bold", 16)
    c.drawString(margin, page_height - margin, "H-Beam Cutting Optimization Report")
    
    c.setFont("Helvetica", 10)
    y = page_height - margin - 30
    
    total_weight = sum(mp.weight for mp in master_plates)
    avg_usage = sum(mp.usage_rate for mp in master_plates) / len(master_plates) * 100
    
    c.drawString(margin, y, f"Total Master Plates: {len(master_plates)}")
    y -= 15
    c.drawString(margin, y, f"Unique Patterns: {len(unique_patterns)}")
    y -= 15
    c.drawString(margin, y, f"Total Weight: {total_weight/1000:,.1f} ton")
    y -= 15
    c.drawString(margin, y, f"Average Usage: {avg_usage:.1f}%")
    y -= 30
    
    # 색상 범례
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margin, y, "Color Legend (Material / Thickness):")
    y -= 15
    
    c.setFont("Helvetica", 9)
    legend_x = margin
    for (mat, thk), color in sorted(material_thickness_colors.items()):
        c.setFillColor(color)
        c.rect(legend_x, y - 3, 15, 12, fill=True, stroke=True)
        c.setFillColor(colors.black)
        c.drawString(legend_x + 20, y, f"{mat} / {thk}mm")
        y -= 15
        if y < 50:
            legend_x += 200
            y = page_height - margin - 100
    
    c.showPage()
    
    # 레이아웃 페이지
    plates_per_page = 3
    draw_height = (page_height - margin * 2 - 30) / plates_per_page
    draw_width = page_width - margin * 2
    
    total_pages = (len(unique_patterns) + plates_per_page - 1) // plates_per_page
    
    for page_idx in range(total_pages):
        c.setFont("Helvetica-Bold", 12)
        c.drawString(margin, page_height - 20, f"Layout Drawings - Page {page_idx + 1}/{total_pages}")
        
        start_idx = page_idx * plates_per_page
        end_idx = min(start_idx + plates_per_page, len(unique_patterns))
        
        for i, (mp, start_seq, end_seq, count) in enumerate(unique_patterns[start_idx:end_idx]):
            y_pos = page_height - margin - 30 - i * draw_height
            plate_color = material_thickness_colors.get((mp.material, mp.thickness), colors.lightgrey)
            
            draw_master_plate_pdf(c, mp, margin, y_pos - draw_height + 20, 
                                 draw_width, draw_height - 30,
                                 start_seq, end_seq, count, plate_color)
        
        c.showPage()
    
    c.save()
    output.seek(0)
    return output


def draw_master_plate_pdf(c, mp, x, y, max_width, max_height, start_seq, end_seq, count, plate_color):
    if start_seq == end_seq:
        title = f"#{start_seq}"
    else:
        title = f"#{start_seq}~{end_seq}"
    
    title += f" | {mp.material} {mp.thickness}mm | {mp.length:.0f}L x {mp.width:.0f}W"
    title += f" | {mp.weight:,.0f}kg | {mp.plate_count}pcs | {mp.usage_rate*100:.0f}%"
    
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x, y + max_height + 5, title)
    
    if count > 1:
        c.setFillColor(colors.red)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(x + max_width - 30, y + max_height - 10, f"x{count}")
        c.setFillColor(colors.black)
    
    margin = 50
    draw_width = max_width - margin * 2
    draw_height = max_height - margin - 10
    
    scale_x = draw_width / mp.length if mp.length > 0 else 1
    scale_y = draw_height / mp.width if mp.width > 0 else 1
    scale = min(scale_x, scale_y) * 0.9
    
    plate_draw_width = mp.length * scale
    plate_draw_height = mp.width * scale
    
    offset_x = x + margin + (draw_width - plate_draw_width) / 2
    offset_y = y + margin / 2 + (draw_height - plate_draw_height) / 2
    
    c.setFillColor(colors.Color(1.0, 1.0, 0.85))
    c.setStrokeColor(colors.black)
    c.setLineWidth(1.5)
    c.rect(offset_x, offset_y, plate_draw_width, plate_draw_height, fill=True, stroke=True)
    
    for plate in mp.placed_plates:
        plate_x_pos = plate.y if plate.y >= 0 and plate.y <= mp.length else 0
        plate_y_pos = plate.x if plate.x >= 0 and plate.x <= mp.width else 0
        
        draw_px = offset_x + plate_x_pos * scale
        draw_py = offset_y + plate_y_pos * scale
        draw_pw = min(plate.length * scale, plate_draw_width - (draw_px - offset_x))
        draw_ph = min(plate.width * scale, plate_draw_height - (draw_py - offset_y))
        
        if draw_pw > 0 and draw_ph > 0:
            c.setFillColor(plate_color)
            c.setStrokeColor(colors.Color(0.3, 0.3, 0.3))
            c.setLineWidth(0.5)
            c.rect(draw_px, draw_py, draw_pw, draw_ph, fill=True, stroke=True)
            
            if draw_pw > 30 and draw_ph > 20:
                c.setFillColor(colors.black)
                c.setFont("Helvetica", 5)
                text_x = draw_px + draw_pw / 2
                text_y = draw_py + draw_ph / 2
                c.drawCentredString(text_x, text_y + 6, f"#{plate.source_no}")
                c.drawCentredString(text_x, text_y, get_plate_type_name(plate.plate_type))
                c.drawCentredString(text_x, text_y - 6, f"{plate.length:.0f}x{plate.width:.0f}")
    
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.black)
    c.drawCentredString(offset_x + plate_draw_width / 2, offset_y - 20, "Length (mm)")
    
    c.saveState()
    c.translate(offset_x - 20, offset_y + plate_draw_height / 2)
    c.rotate(90)
    c.drawCentredString(0, 0, "Width (mm)")
    c.restoreState()
    
    c.setFont("Helvetica", 6)
    c.setStrokeColor(colors.black)
    c.setLineWidth(0.5)
    
    length_step = get_tick_step(mp.length)
    for i in range(0, int(mp.length) + 1, length_step):
        tick_x = offset_x + i * scale
        if tick_x <= offset_x + plate_draw_width + 1:
            c.line(tick_x, offset_y, tick_x, offset_y - 5)
            c.drawCentredString(tick_x, offset_y - 12, str(i))
    
    width_step = get_tick_step(mp.width)
    for i in range(0, int(mp.width) + 1, width_step):
        tick_y = offset_y + i * scale
        if tick_y <= offset_y + plate_draw_height + 1:
            c.line(offset_x, tick_y, offset_x - 5, tick_y)
            c.drawRightString(offset_x - 7, tick_y - 2, str(i))


# ============================================================
# Streamlit UI
# ============================================================
st.set_page_config(page_title="H-Beam 절단 최적화", page_icon="🔧", layout="wide")

st.title("🔧 H-Beam 철판 절단 최적화 시스템")
st.markdown("H-Beam 주문 데이터(CSV/Excel)를 업로드하면 최적의 통철판 절단 계획을 생성합니다.")

with st.sidebar:
    st.header("⚙️ 제약조건")
    st.text(f"길이: {MIN_LENGTH:,} ~ {MAX_LENGTH:,} mm")
    st.text(f"너비: {MIN_WIDTH:,} ~ {MAX_WIDTH:,} mm")
    st.text(f"중량: ≤ {MAX_WEIGHT:,} kg")
    st.markdown("---")
    st.markdown("**선호 범위**")
    st.text(f"길이: {PREFERRED_LENGTH_MIN:,} ~ {PREFERRED_LENGTH_MAX:,} mm")
    st.text(f"너비: {PREFERRED_WIDTH_MIN:,} ~ {PREFERRED_WIDTH_MAX:,} mm")

uploaded = st.file_uploader("📤 주문 파일 업로드 (CSV/Excel)", type=['csv', 'xlsx'])

if uploaded:
    try:
        df = pd.read_csv(uploaded) if uploaded.name.endswith('.csv') else pd.read_excel(uploaded)
        st.success(f"✅ {len(df)}개 주문 로드됨")
        
        with st.expander("📋 데이터 미리보기", expanded=True):
            st.dataframe(df.head(10), use_container_width=True)
        
        st.subheader("📋 컬럼 매핑")
        required = ['no', 'material', 'H', 'B', 'tw', 'tf', 'length', 'quantity']
        col_map = {}
        
        cols = st.columns(4)
        for i, req in enumerate(required):
            with cols[i % 4]:
                default_idx = 0
                for j, col in enumerate(df.columns):
                    if req.lower() in col.lower():
                        default_idx = j
                        break
                col_map[req] = st.selectbox(f"{req}", df.columns.tolist(), index=default_idx, key=f"col_{req}")
        
        if st.button("🚀 최적화 실행", type="primary", use_container_width=True):
            with st.spinner("최적화 중..."):
                orders = []
                for _, row in df.iterrows():
                    try:
                        orders.append({
                            'no': int(row[col_map['no']]),
                            'material': str(row[col_map['material']]),
                            'H': float(row[col_map['H']]),
                            'B': float(row[col_map['B']]),
                            'tw': float(row[col_map['tw']]),
                            'tf': float(row[col_map['tf']]),
                            'length': float(row[col_map['length']]),
                            'quantity': int(row[col_map['quantity']])
                        })
                    except:
                        continue
                
                if not orders:
                    st.error("❌ 유효한 주문 없음")
                    st.stop()
                
                plates = convert_hbeam_to_plates(orders)
                results = optimize_cutting(plates)
                
                st.success(f"✅ 완료! {len(results)}개 통철판")
                
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("통철판", f"{len(results):,}장")
                c2.metric("총 중량", f"{sum(m.weight for m in results)/1000:,.1f}톤")
                c3.metric("평균 사용률", f"{sum(m.usage_rate for m in results)/len(results)*100:.1f}%")
                c4.metric("저효율(<85%)", f"{sum(1 for m in results if m.usage_rate < 0.85)}개")
                
                # 결과 저장 (session state)
                st.session_state['results'] = results
                st.session_state['orders'] = orders
                
                # 테이블
                result_data = [{'No': m.seq, 'Material': m.material, 'Thickness': m.thickness,
                    'Length': m.length, 'Width': m.width, 'Weight': round(m.weight,1),
                    'Plates': m.plate_count, 'Usage%': round(m.usage_rate*100,1)} for m in results]
                st.dataframe(pd.DataFrame(result_data), use_container_width=True, height=400)
                
                # 다운로드 버튼들
                st.subheader("📥 결과 다운로드")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.download_button("📥 통철판 목록 (Excel)", 
                                      create_summary_excel(results).getvalue(),
                                      "master_plates.xlsx",
                                      use_container_width=True)
                
                with col2:
                    st.download_button("📥 상세 결과 (Excel)", 
                                      save_result_formatted(results, orders).getvalue(),
                                      "cutting_result.xlsx",
                                      use_container_width=True)
                
                with col3:
                    st.download_button("📥 레이아웃 도면 (PDF)", 
                                      generate_pdf(results).getvalue(),
                                      "cutting_layout.pdf",
                                      mime="application/pdf",
                                      use_container_width=True)
    
    except Exception as e:
        st.error(f"❌ 오류: {e}")
        st.exception(e)

else:
    st.info("👆 파일을 업로드하세요")
    
    sample = pd.DataFrame({
        'NO': [1, 2, 3], 'Material': ['SS400', 'SM490', 'KS SM355A'],
        'H': [400, 500, 350], 'B': [200, 250, 175],
        'tw': [10, 12, 10], 'tf': [16, 20, 14],
        'Length': [8000, 10000, 7500], 'Quantity': [50, 30, 60]
    })
    st.dataframe(sample, use_container_width=True)
    st.download_button("📥 샘플 CSV", sample.to_csv(index=False), "sample.csv", "text/csv")

st.markdown("---")
st.caption("H-Beam 절단 최적화 v2.0 | 우선순위: 절대규격 > 로스최소화 > 선호규격")
